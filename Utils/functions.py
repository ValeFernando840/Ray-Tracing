import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from scipy.interpolate import make_interp_spline
from scipy.interpolate import PchipInterpolator
from Utils import geo_conversions as gc

def generate_dataframe( latitude_position_tx, longitude_position_tx,elevation_position_tx,
  fc,elev,azim,anio,mmdd,UTI,hora,retardo,rango_terrestre,rango_slant,
  lat_final,lon_final,alt_final,latitudes,longitudes,elevations):
  """"
  Genera un DataFrame con los datos proporcionados y los datos de elevación interpolados realizados aquí.
  Args:
    latitude_position_tx (float): Latitud de la posición del transmisor.
    longitude_position_tx (float): Longitud de la posición del transmisor.
    elevation_position_tx (float): Elevación de la posición del transmisor.
    fc (float): Frecuencia portadora.
    elev (float): Ángulo de elevación.
    azim (float): Ángulo de azimut.
    anio (int): Año.
    mmdd (str): Mes y día en formato MMDD.
    UTI (int): Tiempo Universal Coordinado.
    hora (int): Hora del día.
    retardo (float): Retardo en segundos.
    rango_terrestre (float): Rango terrestre en metros.
    rango_slant (float): Rango oblicuo en metros.
    lat_final (float): Latitud final del receptor.
    lon_final (float): Longitud final del receptor.
    alt_final (float): Altitud final del receptor.
    latitudes (np.array): Array de latitudes para el perfil de elevación.
    longitudes (np.array): Array de longitudes para el perfil de elevación.
    elevations (np.array): Array de elevaciones correspondientes a las latitudes y longitudes.
  Returns:
    pd.DataFrame: DataFrame que contiene todos los datos proporcionados y los datos de elev
  """
  latitudes = latitudes.flatten()
  longitudes = longitudes.flatten()
  elevations = elevations.flatten()
  lat_filt, long_filt, elev_filt = filter_on_elevations(latitudes,longitudes,elevations)
  latitudes ,longitudes,elevations = filter_unique_coordinates(lat_filt,long_filt,elev_filt)
  x,y,z = gc.transform_coords_cartesian(latitudes, longitudes, elevations)
  x_int, y_int, z_int = interpolate_with_Pchip(x,y,z)
  phi_int, theta_int, rho_int = gc.transform_cartesian_to_spherical(x_int,y_int,z_int)

    # Create a Data Frame
  data = {
      "latitude_pos_tx": latitude_position_tx,
      "longitude_pos_tx": longitude_position_tx,
      "elevation_pos_tx": elevation_position_tx,
      "fc": fc,
      "elevation": elev,
      "azimuth": azim,
      "year": anio,
      "mmdd": mmdd,
      "UTI": UTI,
      "hour": hora,
      "delay": retardo,
      "terrestrial_range": rango_terrestre,
      "slant_range": rango_slant,
      "final_latitude": lat_final,
      "final_longitude": lon_final,
      "final_elevation": alt_final,
      "latitudes": [latitudes],
      "longitudes": [longitudes],
      "alturas": [elevations]
    }
    #expand arrays in columns sepatare
  # for i in range(len(latitudes)):
  #   data[f'lat_{i+1}'] = latitudes[i]
  # for i in range(len(longitudes)):
  #   data[f'long_{i+1}'] = longitudes[i]
  # for i in range(len(elevations)):
  #   data[f'elev_{i+1}'] = elevations[i]

  df = pd.DataFrame(data)
  return df

# La funcion agrega al dataset una nueva linea
def add_to_dataset(df):
    address = os.getcwd()
    new = "dataset"
    new_address = os.path.join(address, new)
    df.to_csv(new_address+"/dataset.csv", index=False, header = False, mode = "a")
    return
def add_to_excel(dir, line_df):
  """
  Agrega una nueva línea a un archivo Excel existente.
  Args:
      dir (str): Ruta del archivo Excel.
      line_df (pd.DataFrame): DataFrame que contiene la línea a agregar.
  Returns:
      None
  """
  if not os.path.exists(dir):
      raise FileNotFoundError(f"El archivo {dir} no existe.")
  
  if line_df.empty:
      raise ValueError("El DataFrame proporcionado está vacío.")

  # Cargar el libro para saber en qué fila termina
  book = load_workbook(dir)
  hoja = book.active
  last_row = hoja.max_row

  # Escribir en modo append
  with pd.ExcelWriter(dir, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
      line_df.to_excel(writer, index=False, header=False, startrow=last_row)
  return
# La función genera un link con las ubicaciones en el mapa de los 2 puntos
# proporcionados
def coordinates_on_map( initial_latitude, initial_longitude, 
                      final_latitude, final_longitude):
    """
    Genera un enlace de Google Maps para las coordenadas proporcionadas.
    Args:
        initial_latitude (float): Latitud inicial.
        initial_longitude (float): Longitud inicial.
        final_latitude (float): Latitud final.
        final_longitude (float): Longitud final.
    Returns:
        None: Imprime la URL de Google Maps.
    """
    final_latitude = final_latitude[0]
    final_longitude = final_longitude[0]
    url = f"https://www.google.com/maps/dir/?api=1&origin={initial_latitude},{initial_longitude}&destination={final_latitude},{final_longitude}&travelmode=driving"
    print("Open the following URL in your browser:\n", url)
    return

# Esta Función recibirá un conjunto de Lat, Long, y elevs
# y devuelve con conjunto de 100 muestras fijas.
def interpolate3d(latitudes,longitudes,elevations):
  """
  Esta función interpola un conjunto de datos 3D (latitudes, longitudes y elevaciones)
  utilizando splines lineales para generar un conjunto de datos con 100 muestras fijas.
  Args:
    latitudes (np.array): Array de latitudes.
    longitudes (np.array): Array de longitudes.
    elevations (np.array): Array de elevaciones.
  Returns:
    tuple: Tupla que contiene arrays de latitudes, longitudes y elevaciones interpoladas.
  """
  posiciones = np.linspace(0, len(latitudes) - 1, len(latitudes))  # Crear una secuencia de índices para spline
  
  # Crear splines separados para latitudes, longitudes y elevaciones
  spline_lat = make_interp_spline(posiciones, latitudes, k=1)
  spline_lon = make_interp_spline(posiciones, longitudes, k=1)
  spline_elev = make_interp_spline(posiciones, elevations, k=1)
  
  # Generar nuevas posiciones con 100 puntos
  posiciones_nuevas = np.linspace(0, len(latitudes) - 1, 100)
  
  # Generar valores interpolados
  latitudes_int = spline_lat(posiciones_nuevas)
  longitudes_int = spline_lon(posiciones_nuevas)
  elevations_int = spline_elev(posiciones_nuevas)  

  return latitudes_int,longitudes_int,elevations_int

def filter_on_elevations(latitudes,longitudes,elevations):
  """ 
  Esta función filtra las elevaciones negativas de un conjunto de datos.
  Args:
    latitudes (np.array): Array de latitudes.
    longitudes (np.array): Array de longitudes.
    elevations (np.array): Array de elevaciones. 
  Returns:
    tuple: Tupla que contiene arrays de latitudes, longitudes y elevaciones filtradas
  """
  lat_filt = []
  long_filt = []
  elev_filt = []

  for i in range(len(elevations)):
    if elevations[i] >= 0:
       lat_filt.append(latitudes[i])
       long_filt.append(longitudes[i])
       elev_filt.append(elevations[i])
  lat_filt = np.array(lat_filt)
  long_filt = np.array(long_filt)  
  elev_filt = np.array(elev_filt)
  
  #Verificar Longitudes
  if len(lat_filt) != len(long_filt) or len(long_filt) != len(elev_filt):
    raise ValueError("Los arrays latitudes, longitudes y elevaciones deben tener la misma longitud")
  # Verificar valores NaN o Inf
  if np.any(np.isnan(lat_filt)) or np.any(np.isnan(long_filt)) or np.any(np.isnan(elev_filt)):
    raise ValueError("Los arrays contienen valores NaN")
  
  if np.any(np.isinf(lat_filt)) or np.any(np.isinf(long_filt)) or np.any(np.isinf(elev_filt)):
    raise ValueError("Los arrays contienen valores infinitos")
  
  return lat_filt, long_filt, elev_filt

def filter_unique_coordinates(latitudes,longitudes,elevations):
  """
  Esta función filtra las coordenadas únicas de un conjunto de datos.
  Args:
    latitudes (np.array): Array de latitudes.
    longitudes (np.array): Array de longitudes.
    elevations (np.array): Array de elevaciones.
  Returns:
    tuple: Tupla que contiene arrays de latitudes, longitudes y elevaciones únicas.
  """
  data = {
    'latitudes': latitudes,
    'longitudes': longitudes,
    'elevations': elevations 
  }
  df = pd.DataFrame(data)

  df = df.drop_duplicates()
  latitudes = df['latitudes'].to_numpy()
  longitudes = df['longitudes'].to_numpy()
  elevations = df['elevations'].to_numpy() 
  return latitudes, longitudes, elevations

def csv_to_excel():
  """
  Permite convertir un archivo CSV a formato Excel (.xlsx).
  Verifica que el archivo CSV exista en la ruta especificada.
  Verifica la dirección de salida.
  """
  address = os.getcwd()
  new = "dataset"
  new_address = os.path.join(address, new)
  data = pd.read_csv(new_address+"/nuevo.csv")
  print(new_address)
  data.to_excel(new_address+"/nuevo_excel.xlsx",index=False)
  return

def interpolate_with_Pchip(x,y,z):
	"""
	Interpola puntos 3D utilizando PCHIP (Piecewise Cubic Hermite Interpolating Polynomial).
	Args:
		x (_array_): _Array de coordenadas x._
		y (_array_): _Array de coordenadas y._
		z (_array_): _Array de coordenadas z._
	Returns:
		x_pchip (_array_): _Array de coordenadas x interpoladas._	
		y_pchip (_array_): _Array de coordenadas y interpoladas._
		z_pchip (_array_): _Array de coordenadas z interpoladas._
	"""
	dx = np.diff(x)
	dy = np.diff(y)
	dz = np.diff(z)

	dist = np.sqrt(dx**2 + dy**2 + dz**2)
	cumulative_dist = np.concatenate(([0], np.cumsum(dist)))
	t_coarse = cumulative_dist
	t_fine = np.linspace(0, cumulative_dist[-1],200)

	interp_x = PchipInterpolator(t_coarse, x)
	interp_y = PchipInterpolator(t_coarse, y)
	interp_z = PchipInterpolator(t_coarse, z)
	x_pchip = interp_x(t_fine)
	y_pchip = interp_y(t_fine)
	z_pchip = interp_z(t_fine)
	return x_pchip, y_pchip, z_pchip